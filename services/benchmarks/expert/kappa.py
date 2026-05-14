#!/usr/bin/env python3
"""
Cohen's Kappa inter-rater reliability analysis for expert review spreadsheets.

Reads two completed expert-rated .xlsx files (from generate.py), computes
weighted Cohen's Kappa per Likert dimension, and generates a Markdown report.

Usage:
  cd services && python -m benchmarks.expert.kappa --rater-a path/a.xlsx --rater-b path/b.xlsx
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import openpyxl
from sklearn.metrics import cohen_kappa_score

from benchmarks.expert.columns import LIKERT_COLS
from benchmarks.io import RESULTS_DIR

KAPPA_THRESHOLDS = [
    (0.81, 1.01, "Almost Perfect"),
    (0.61, 0.80, "Substantial"),
    (0.41, 0.60, "Moderate"),
    (0.21, 0.40, "Fair"),
    (0.00, 0.20, "Slight"),
    (-1.0, 0.00, "Poor"),
]


def read_ratings(xlsx_path: str | Path) -> dict[int, dict[str, int]]:
    """Read Likert ratings from an expert review spreadsheet.

    Returns dict keyed by 0-based row index mapping to rating dicts.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb["Expert Review"]

    # Find column indices for Likert columns from header row
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_indices = {}
    for col_name in LIKERT_COLS:
        if col_name in header:
            col_indices[col_name] = header.index(col_name)

    ratings = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row_ratings = {}
        skip = False
        for col_name, col_pos in col_indices.items():
            val = row[col_pos] if col_pos < len(row) else None
            if val is None:
                skip = True
                break
            try:
                row_ratings[col_name] = int(val)
            except (ValueError, TypeError):
                skip = True
                break
        if not skip and row_ratings:
            ratings[row_idx] = row_ratings

    wb.close()
    return ratings


def compute_kappa(
    ratings_a: dict[int, dict[str, int]],
    ratings_b: dict[int, dict[str, int]],
    dimension: str,
    weights: str = "linear",
) -> float | None:
    """Compute Cohen's Kappa for a single dimension on common rows."""
    common = sorted(set(ratings_a.keys()) & set(ratings_b.keys()))
    if len(common) < 2:
        return None
    y1 = [ratings_a[i][dimension] for i in common]
    y2 = [ratings_b[i][dimension] for i in common]
    return cohen_kappa_score(y1, y2, weights=weights)


def compute_raw_agreement(
    ratings_a: dict[int, dict[str, int]],
    ratings_b: dict[int, dict[str, int]],
    dimension: str,
) -> float | None:
    """Compute fraction of exact matches on common rows."""
    common = sorted(set(ratings_a.keys()) & set(ratings_b.keys()))
    if not common:
        return None
    matches = sum(1 for i in common if ratings_a[i][dimension] == ratings_b[i][dimension])
    return matches / len(common)


def interpret_kappa(value: float | None) -> str:
    """Map a kappa value to its Landis-Koch interpretation label."""
    if value is None:
        return "N/A (insufficient data)"
    for low, high, label in KAPPA_THRESHOLDS:
        if low <= value <= high:
            return label
    return "Unknown"


def analyse(rater_a_path: Path, rater_b_path: Path) -> dict:
    """Run full inter-rater reliability analysis on two completed spreadsheets."""
    ratings_a = read_ratings(rater_a_path)
    ratings_b = read_ratings(rater_b_path)
    common = set(ratings_a.keys()) & set(ratings_b.keys())

    dimensions = {}
    for dim in LIKERT_COLS:
        kappa = compute_kappa(ratings_a, ratings_b, dim, weights="linear")
        raw = compute_raw_agreement(ratings_a, ratings_b, dim)
        dimensions[dim] = {
            "kappa": kappa,
            "raw_agreement": raw,
            "interpretation": interpret_kappa(kappa),
        }

    return {
        "meta": {
            "rater_a": str(rater_a_path),
            "rater_b": str(rater_b_path),
            "common_rows": len(common),
            "rater_a_rows": len(ratings_a),
            "rater_b_rows": len(ratings_b),
        },
        "dimensions": dimensions,
    }


def generate_report(results: dict) -> str:
    """Generate a Markdown inter-rater reliability report."""
    meta = results["meta"]
    dims = results["dimensions"]
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    lines = [
        "# Expert Validation — Inter-Rater Reliability Report",
        "",
        "## Overview",
        "",
        f"- **Date:** {now}",
        f"- **Rater A:** {Path(meta['rater_a']).name}",
        f"- **Rater B:** {Path(meta['rater_b']).name}",
        f"- **Rater A rows:** {meta['rater_a_rows']}",
        f"- **Rater B rows:** {meta['rater_b_rows']}",
        f"- **Common rows:** {meta['common_rows']}",
        "",
        "## Cohen's Kappa — Overall",
        "",
        "| Dimension | Kappa | Raw Agreement | Interpretation |",
        "|-----------|-------|---------------|----------------|",
    ]

    for dim, data in dims.items():
        kappa_str = f"{data['kappa']:.3f}" if data["kappa"] is not None else "N/A"
        raw_str = f"{data['raw_agreement']:.1%}" if data["raw_agreement"] is not None else "N/A"
        lines.append(f"| {dim} | {kappa_str} | {raw_str} | {data['interpretation']} |")

    lines.extend([
        "",
        "## Interpretation Guide",
        "",
        "| Kappa Range | Interpretation (Landis & Koch, 1977) |",
        "|-------------|--------------------------------------|",
    ])
    for low, high, label in KAPPA_THRESHOLDS:
        if low < 0:
            lines.append(f"| < 0.00 | {label} |")
        else:
            lines.append(f"| {low:.2f} - {high:.2f} | {label} |")

    lines.extend([
        "",
        "## Notes",
        "",
        "- Weighted Cohen's Kappa (linear weights) used for ordinal Likert scale data.",
        "- Raw agreement reported alongside Kappa to account for the Kappa paradox "
        "(high agreement with skewed distributions can yield low Kappa).",
        "",
    ])

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Compute Cohen's Kappa inter-rater reliability from expert review spreadsheets"
    )
    parser.add_argument("--rater-a", required=True, help="Path to first expert's completed XLSX")
    parser.add_argument("--rater-b", required=True, help="Path to second expert's completed XLSX")
    parser.add_argument("--output", default=None, help="Output path for report")
    args = parser.parse_args()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_path = args.output or str(RESULTS_DIR / f"expert_kappa_{ts}.md")

    results = analyse(Path(args.rater_a), Path(args.rater_b))
    report = generate_report(results)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    Path(output_path).write_text(report, encoding="utf-8")

    # Print summary to stdout
    print()
    for line in report.split("\n"):
        if line.startswith("|") or line.startswith("##"):
            print(line)
    print()
    print(f"Report written to {output_path}")


if __name__ == "__main__":
    main()
