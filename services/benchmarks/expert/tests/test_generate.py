"""Tests for the expert review spreadsheet generator."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import openpyxl
import pytest

from benchmarks.expert.columns import ALL_COLS, LIKERT_COLS, QUESTION_COLS, RATING_COLS
from benchmarks.expert.generate import build_rows, write_xlsx


def _make_questions(n=3):
    """Create mock question dicts matching generate_qg output shape."""
    questions = []
    for i in range(n):
        questions.append({
            "id": f"q-test-abc123-mcq-{i+1}",
            "kind": "mcq",
            "text": f"What is concept {i+1}?",
            "options": [f"Option A{i}", f"Option B{i}", f"Option C{i}", f"Option D{i}"],
            "correct": "A",
            "why": "Because it is correct.",
            "keywords": [f"keyword_{i}"],
            "source_spans": [{"doc": "test.pdf", "text": f"Source text for question {i+1}. " * 5}],
        })
    return questions


class TestBuildRows:
    def test_converts_to_row_dicts(self):
        questions = _make_questions(2)
        rows = build_rows(questions, "beginner")
        assert len(rows) == 2
        for row in rows:
            assert set(row.keys()) == set(ALL_COLS)
            assert row["difficulty_level"] == "beginner"

    def test_rating_cols_empty(self):
        rows = build_rows(_make_questions(1), "beginner")
        for col in RATING_COLS:
            assert rows[0][col] == ""

    def test_serializes_options(self):
        rows = build_rows(_make_questions(1), "intermediate")
        opts = rows[0]["options"]
        assert opts.startswith("A) ")
        assert "; B) " in opts
        assert "; C) " in opts
        assert "; D) " in opts

    def test_truncates_source_snippet(self):
        q = _make_questions(1)
        q[0]["source_spans"] = [{"doc": "test.pdf", "text": "X" * 300}]
        rows = build_rows(q, "advanced")
        assert len(rows[0]["source_snippet"]) == 200

    def test_empty_source_spans(self):
        q = _make_questions(1)
        q[0]["source_spans"] = []
        rows = build_rows(q, "beginner")
        assert rows[0]["source_snippet"] == ""

    def test_empty_keywords(self):
        q = _make_questions(1)
        q[0]["keywords"] = []
        rows = build_rows(q, "beginner")
        assert rows[0]["keyword"] == ""


class TestWriteXlsx:
    def test_creates_valid_xlsx(self, tmp_path):
        rows = build_rows(_make_questions(3), "beginner")
        out = tmp_path / "test.xlsx"
        result = write_xlsx(rows, out)
        assert result == out
        assert out.exists()

    def test_two_sheets(self, tmp_path):
        rows = build_rows(_make_questions(2), "beginner")
        out = tmp_path / "test.xlsx"
        write_xlsx(rows, out)
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["Instructions", "Expert Review"]

    def test_header_row_matches_all_cols(self, tmp_path):
        rows = build_rows(_make_questions(2), "beginner")
        out = tmp_path / "test.xlsx"
        write_xlsx(rows, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Expert Review"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, len(ALL_COLS) + 1)]
        assert headers == ALL_COLS

    def test_data_validation_on_likert_cols(self, tmp_path):
        rows = build_rows(_make_questions(3), "beginner")
        out = tmp_path / "test.xlsx"
        write_xlsx(rows, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Expert Review"]
        dvs = ws.data_validations.dataValidation
        assert len(dvs) >= 1
        # Check that validation formula contains 1-5
        for dv in dvs:
            assert "1,2,3,4,5" in (dv.formula1 or "")

    def test_question_cells_locked(self, tmp_path):
        rows = build_rows(_make_questions(2), "beginner")
        out = tmp_path / "test.xlsx"
        write_xlsx(rows, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Expert Review"]
        # Check first question column cell in row 2
        cell = ws.cell(row=2, column=1)  # question_text
        assert cell.protection.locked is True

    def test_rating_cells_unlocked(self, tmp_path):
        rows = build_rows(_make_questions(2), "beginner")
        out = tmp_path / "test.xlsx"
        write_xlsx(rows, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Expert Review"]
        # Check first Likert column cell in row 2
        likert_col_idx = ALL_COLS.index("factual_correctness") + 1
        cell = ws.cell(row=2, column=likert_col_idx)
        assert cell.protection.locked is False

    def test_sheet_protection_enabled(self, tmp_path):
        rows = build_rows(_make_questions(2), "beginner")
        out = tmp_path / "test.xlsx"
        write_xlsx(rows, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Expert Review"]
        assert ws.protection.sheet is True

    def test_instructions_content(self, tmp_path):
        rows = build_rows(_make_questions(1), "beginner")
        out = tmp_path / "test.xlsx"
        write_xlsx(rows, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Instructions"]
        assert ws["A1"].value == "Expert Review Instructions"
        assert "Rating Scale" in (ws["A5"].value or "")
        assert "1 = Completely wrong" in (ws["A6"].value or "")
        assert "keyword" in (ws["A13"].value or "").lower()

    def test_freeze_panes(self, tmp_path):
        rows = build_rows(_make_questions(2), "beginner")
        out = tmp_path / "test.xlsx"
        write_xlsx(rows, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Expert Review"]
        assert ws.freeze_panes == "A2"
