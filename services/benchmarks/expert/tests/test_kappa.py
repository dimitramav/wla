"""Tests for the Cohen's Kappa inter-rater reliability analysis."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import openpyxl
import pytest

from benchmarks.expert.columns import ALL_COLS, LIKERT_COLS
from benchmarks.expert.kappa import (
    compute_kappa,
    compute_raw_agreement,
    generate_report,
    interpret_kappa,
    read_ratings,
)


def create_rated_xlsx(tmp_path: Path, name: str, ratings_data: list[dict]) -> Path:
    """Create a minimal XLSX with Expert Review sheet and Likert values filled in."""
    path = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expert Review"

    # Header row
    for col_idx, col_name in enumerate(ALL_COLS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, row_data in enumerate(ratings_data, start=2):
        for col_idx, col_name in enumerate(ALL_COLS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    wb.save(str(path))
    return path


def _make_rated_row(fc=4, pa=4, sf=4):
    """Create a row dict with given Likert ratings."""
    return {
        "question_text": "What is X?",
        "correct_answer": "A",
        "options": "A) X; B) Y; C) Z; D) W",
        "explanation": "Because X is correct.",
        "source_document": "test.pdf",
        "keyword": "test_kw",
        "difficulty_level": "beginner",
        "factual_correctness": fc,
        "pedagogical_alignment": pa,
        "source_fidelity": sf,
        "rationale": "",
    }


class TestReadRatings:
    def test_reads_valid_ratings(self, tmp_path):
        rows = [_make_rated_row(5, 4, 3), _make_rated_row(4, 5, 4)]
        path = create_rated_xlsx(tmp_path, "rater.xlsx", rows)
        ratings = read_ratings(path)
        assert len(ratings) == 2
        assert ratings[0]["factual_correctness"] == 5
        assert ratings[1]["source_fidelity"] == 4

    def test_skips_rows_with_missing_ratings(self, tmp_path):
        rows = [_make_rated_row(5, 4, 3), _make_rated_row(None, 5, 4)]
        path = create_rated_xlsx(tmp_path, "rater.xlsx", rows)
        ratings = read_ratings(path)
        assert len(ratings) == 1
        assert 0 in ratings


class TestComputeKappa:
    def test_identical_ratings_returns_one(self):
        ratings_a = {0: {"factual_correctness": 5}, 1: {"factual_correctness": 4}, 2: {"factual_correctness": 3}}
        ratings_b = {0: {"factual_correctness": 5}, 1: {"factual_correctness": 4}, 2: {"factual_correctness": 3}}
        kappa = compute_kappa(ratings_a, ratings_b, "factual_correctness")
        assert kappa == pytest.approx(1.0)

    def test_common_rows_only(self):
        ratings_a = {0: {"factual_correctness": 5}, 1: {"factual_correctness": 4}, 2: {"factual_correctness": 3}}
        ratings_b = {1: {"factual_correctness": 4}, 2: {"factual_correctness": 3}, 3: {"factual_correctness": 2}}
        kappa = compute_kappa(ratings_a, ratings_b, "factual_correctness")
        # Only rows 1 and 2 are common, with identical values
        assert kappa == pytest.approx(1.0)

    def test_insufficient_data_returns_none(self):
        ratings_a = {0: {"factual_correctness": 5}}
        ratings_b = {0: {"factual_correctness": 5}}
        # Only 1 common row, need at least 2
        kappa = compute_kappa(ratings_a, ratings_b, "factual_correctness")
        assert kappa is None

    def test_no_overlap_returns_none(self):
        ratings_a = {0: {"factual_correctness": 5}}
        ratings_b = {1: {"factual_correctness": 5}}
        kappa = compute_kappa(ratings_a, ratings_b, "factual_correctness")
        assert kappa is None


class TestComputeRawAgreement:
    def test_perfect_agreement(self):
        ratings_a = {0: {"factual_correctness": 5}, 1: {"factual_correctness": 4}}
        ratings_b = {0: {"factual_correctness": 5}, 1: {"factual_correctness": 4}}
        raw = compute_raw_agreement(ratings_a, ratings_b, "factual_correctness")
        assert raw == pytest.approx(1.0)

    def test_partial_agreement(self):
        ratings_a = {0: {"factual_correctness": 5}, 1: {"factual_correctness": 4}}
        ratings_b = {0: {"factual_correctness": 5}, 1: {"factual_correctness": 3}}
        raw = compute_raw_agreement(ratings_a, ratings_b, "factual_correctness")
        assert raw == pytest.approx(0.5)

    def test_no_common_rows(self):
        ratings_a = {0: {"factual_correctness": 5}}
        ratings_b = {1: {"factual_correctness": 5}}
        raw = compute_raw_agreement(ratings_a, ratings_b, "factual_correctness")
        assert raw is None


class TestInterpretKappa:
    @pytest.mark.parametrize("value,expected", [
        (0.85, "Almost Perfect"),
        (0.75, "Substantial"),
        (0.5, "Moderate"),
        (0.3, "Fair"),
        (0.1, "Slight"),
        (-0.1, "Poor"),
        (None, "N/A (insufficient data)"),
    ])
    def test_interpret(self, value, expected):
        assert interpret_kappa(value) == expected


class TestGenerateReport:
    def test_contains_required_sections(self):
        results = {
            "meta": {
                "rater_a": "rater_a.xlsx",
                "rater_b": "rater_b.xlsx",
                "common_rows": 10,
                "rater_a_rows": 10,
                "rater_b_rows": 10,
            },
            "dimensions": {
                "factual_correctness": {"kappa": 0.75, "raw_agreement": 0.8, "interpretation": "Substantial"},
                "pedagogical_alignment": {"kappa": 0.6, "raw_agreement": 0.7, "interpretation": "Moderate"},
                "source_fidelity": {"kappa": None, "raw_agreement": None, "interpretation": "N/A (insufficient data)"},
            },
        }
        report = generate_report(results)
        assert "## Cohen's Kappa" in report
        assert "| Dimension |" in report
        assert "| factual_correctness |" in report
        assert "0.750" in report
        assert "80.0%" in report
        assert "Landis & Koch" in report
        assert "linear weights" in report
