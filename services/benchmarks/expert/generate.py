#!/usr/bin/env python3
"""
Expert review spreadsheet generator for Watch-Listen-Act Phase 10.

Generates quiz questions via the RAG pipeline and exports them to an .xlsx
spreadsheet with embedded instructions, locked question columns, and unlocked
Likert-scale rating columns with dropdown validation.

Usage:
  cd services && python -m benchmarks.expert.generate
  cd services && python -m benchmarks.expert.generate --n-per-level 5 --n-sets 1
"""

import argparse
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection

from benchmarks.config import TOPIC
from benchmarks.expert.columns import (
    ALL_COLS,
    LEVELS,
    LIKERT_COLS,
    LIKERT_LABELS,
    N_PER_LEVEL,
    N_SETS,
    QUESTION_COLS,
    RATING_COLS,
)
from benchmarks.io import RESULTS_DIR


def build_rows(questions: list[dict], level: str) -> list[dict]:
    """Convert generate_qg question dicts into row dicts keyed by ALL_COLS."""
    rows = []
    for q in questions:
        options_list = q.get("options", [])
        # Strip any existing letter prefix (e.g. "A) ") the LLM already added
        cleaned = [re.sub(r'^[A-Da-d]\)\s*', '', opt) for opt in options_list[:4]]
        letters = ["A", "B", "C", "D"]
        serialized = "; ".join(
            f"{letters[i]}) {opt}" for i, opt in enumerate(cleaned)
        )

        source_spans = q.get("source_spans", [])
        source_doc = ""
        if source_spans and source_spans[0].get("doc"):
            source_doc = source_spans[0]["doc"]

        keywords = q.get("keywords", [])
        keyword = keywords[0] if keywords else ""

        row = {
            "question_text": q.get("text", ""),
            "correct_answer": q.get("correct", ""),
            "options": serialized,
            "explanation": q.get("why", ""),
            "source_document": source_doc,
            "keyword": keyword,
            "difficulty_level": level,
        }
        for col in RATING_COLS:
            row[col] = ""
        rows.append(row)
    return rows


def build_instructions_sheet(ws):
    """Write instructions content to the given worksheet."""
    ws.merge_cells("A1:D1")
    ws["A1"] = "Expert Review Instructions"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = (
        "Purpose: Rate the quality of AI-generated quiz questions for teacher "
        "training on student mental health."
    )

    ws["A5"] = "Rating Scale (1-5):"
    ws["A5"].font = Font(bold=True)
    for score, label in LIKERT_LABELS.items():
        ws[f"A{5 + score}"] = f"{score} = {label}"

    ws["A12"] = "Question Structure:"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = (
        "Each question targets a specific keyword from the topic's taxonomy. "
        "The 'keyword' column shows which concept the question was generated for."
    )
    ws["A14"] = (
        "Questions are distributed evenly: 2 questions per keyword, 5 keywords per "
        "difficulty level, across 3 levels (beginner, intermediate, advanced)."
    )

    ws["A16"] = "Criteria Definitions:"
    ws["A16"].font = Font(bold=True)
    ws["A17"] = (
        "factual_correctness: Is the question and its correct answer factually accurate?"
    )
    ws["A18"] = (
        "pedagogical_alignment: Is the question appropriate for training teachers on this topic?"
    )
    ws["A19"] = (
        "source_fidelity: Does the question faithfully reflect the source document cited?"
    )

    ws["A21"] = (
        "rationale: Optional free-text explanation for your ratings, "
        "especially for scores \u2264 3."
    )

    ws["A23"] = (
        "Estimated time: ~1-2 minutes per question, ~2 hours total for 60 questions."
    )
    ws["A23"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 80


COL_WIDTHS = {
    "question_text": 50,
    "correct_answer": 15,
    "options": 40,
    "explanation": 50,
    "source_document": 40,
    "keyword": 20,
    "difficulty_level": 15,
    "factual_correctness": 20,
    "pedagogical_alignment": 20,
    "source_fidelity": 20,
    "rationale": 40,
}


def write_xlsx(rows: list[dict], output_path: str | Path) -> Path:
    """Create the expert review workbook and save to output_path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Instructions
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    build_instructions_sheet(ws_instr)

    # Sheet 2: Expert Review (set as active)
    ws = wb.create_sheet("Expert Review")
    wb.active = wb.sheetnames.index("Expert Review")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)

    # Header row
    for col_idx, col_name in enumerate(ALL_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(ALL_COLS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            if col_name in QUESTION_COLS:
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)

    # Data validation for Likert columns
    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv.error = "Please enter a value between 1 and 5"
    dv.errorTitle = "Invalid rating"
    ws.add_data_validation(dv)

    for col_name in LIKERT_COLS:
        col_idx = ALL_COLS.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{len(rows) + 1}")

    # Column widths
    for col_name, width in COL_WIDTHS.items():
        col_idx = ALL_COLS.index(col_name) + 1
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Sheet protection
    ws.protection = SheetProtection(sheet=True, password="review")

    # Freeze panes
    ws.freeze_panes = "A2"

    wb.save(str(output_path))
    return output_path


def load_level_keywords(topic: str, level: str) -> list[str]:
    """Load keywords for a specific difficulty level from keywords.yaml."""
    import yaml

    kw_path = Path(__file__).parent.parent.parent.parent / "content" / topic / "keywords.yaml"
    with open(kw_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)

    level_map = {"beginner": "1", "intermediate": "2", "advanced": "3"}
    level_key = level_map.get(level, level)
    return data.get(topic, {}).get(level_key, [])


def main():
    parser = argparse.ArgumentParser(
        description="Generate expert review spreadsheet from RAG pipeline"
    )
    parser.add_argument("--topic", default=TOPIC, help="Topic name")
    parser.add_argument("--docset-hash", default=None, help="Docset hash (auto-detected if omitted)")
    parser.add_argument("--n-per-level", type=int, default=N_PER_LEVEL, help="Questions per level per set (distributed evenly across 5 keywords)")
    parser.add_argument("--n-sets", type=int, default=N_SETS, help="Number of question sets")
    parser.add_argument("--seed", type=int, default=42, help="Base seed for reproducibility")
    parser.add_argument("--output", default=None, help="Output path (default: results/expert_<ts>.xlsx)")
    args = parser.parse_args()

    from rag.ingest import ingest_topic
    from rag.qg import generate_qg

    # Ingest to get docset_hash
    docset_hash = args.docset_hash
    if not docset_hash:
        res = ingest_topic(args.topic, force=False)
        docset_hash = res.get("docset_hash", "")
        if not docset_hash:
            print("ERROR: ingest_topic returned no docset_hash")
            sys.exit(1)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_path = args.output or str(RESULTS_DIR / f"expert_{args.topic}_{ts}.xlsx")

    all_rows = []
    n_per_kw = args.n_per_level // 5  # 5 keywords per level → 2 questions each
    total_calls = args.n_sets * len(LEVELS) * 5  # sets × levels × keywords
    call_num = 0

    print(f"Plan: {args.n_sets} set(s) × {len(LEVELS)} levels × 5 keywords × {n_per_kw} q/kw = {args.n_sets * len(LEVELS) * 5 * n_per_kw} questions")
    print(f"Total generate_qg calls: {total_calls}")
    print()

    for set_idx in range(args.n_sets):
        for level in LEVELS:
            keywords = load_level_keywords(args.topic, level)
            difficulty_profile = {"difficulty_label": level}
            # Generate n_per_kw questions per keyword for even coverage
            for kw_idx, kw in enumerate(keywords):
                call_num += 1
                print(f"[{call_num}/{total_calls}] set={set_idx+1} level={level} keyword='{kw}' ({n_per_kw} MCQs)...", flush=True)
                result = generate_qg(
                    topic=args.topic,
                    docset_hash=docset_hash,
                    mix={"mcq": n_per_kw, "yesno": 0},
                    seed=str(args.seed + set_idx * 100 + kw_idx),
                    keywords=[kw],
                    weak_keywords=[],
                    weak_focus_ratio=0.0,
                    difficulty_profile=difficulty_profile,
                )
                questions = result.get("questions", [])
                all_rows.extend(build_rows(questions[:n_per_kw], level))
                print(f"  ✓ Got {len(questions)} questions (total so far: {len(all_rows)})", flush=True)

    write_xlsx(all_rows, output_path)
    print(f"\nGenerated {len(all_rows)} questions -> {output_path}")


if __name__ == "__main__":
    main()
